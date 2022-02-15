import os
import subprocess
import re
import json
from openpyxl import Workbook
import openpyxl


print("What format are you need? Please, write the number. \n" + "1. xlsx \n" + "2. json")
format = input()
while format:
    if format is "1" or format is "2":
        break
    else:
        print("Error format! Please, try again.")
        format = input()



if format is "1":
    wb = Workbook()
    ws = wb.active
    ws.title = "New Title"
    ws['A1'] = "#N"
    ws['B1'] = "Channel"
    ws['C1'] = "Address"
    ws['D1'] = "Resolution"
    ws['E1'] = "V-codec"
    ws['F1'] = "A-codec"
    ws['G1'] = "A-codec 2"
    ws['H1'] = "Subtitle"
    ws['I1'] = "Subtitle 2"
else:
    json_f = open("m9.json", "w")
    json_f.close()



n = 2
i = 0
while i < 255:
    i += 1


    mCast = "http://207.110.52.50:4022/udp/233.166.172." + str(i) + ":1234"



    
    output = (subprocess.run(["docker", "run", "-it", "--rm", "--network", "host", "nfs01.techstudio.tv/ffprobe:latest", "-v", "quiet", "-print_format", "json", "-show_format", "-show_streams", "-i", mCast], stdout=subprocess.PIPE))


    data = json.loads(output.stdout.decode("utf8"))
    if not data:
        continue
    print(data)
    if format is "1":
        ws['A' + str(n)] = str(n-1)
        ws['C' + str(n)] = mCast
        sub_check = 0
        aud_check = 0
        for check in data["streams"]:
            if "codec_type" not in check:
                continue
            elif check["codec_type"] == "video":
                print(check["codec_name"])
                ws['E' + str(n)] = check["codec_name"]
                resolution = str(check["width"]) + "x" + str(check["height"])
            elif check["codec_type"] == "audio":
                if aud_check != 0:
                    print(check["codec_name"])
                    ws['G' + str(n)] = check["codec_name"]
                    aud_check = 0
                else:
                    print(check["codec_name"])
                    ws['F' + str(n)] = check["codec_name"]
                    aud_check += 1
            elif check["codec_type"] == "subtitle":
                if sub_check != 0:                  
                    ws['I' + str(n)] = check["codec_name"]
                    print(check["codec_name"])
                    sub_check = 0
                else:
                    ws['H' + str(n)] = check["codec_name"]
                    print(check["codec_name"])
                    sub_check += 1
            

        if not resolution:
            resolution = "--------"
        ws['D' + str(n)] = resolution
        n += 1
    else:
        json_f = open("m9.json", "a")
        json_f.write(json.dumps(data))
        json_f.close



    
j = 0
while j < 255:
    j += 1

    mCast = "http://207.110.52.50:4022/udp/233.166.173." + str(j) + ":1234"
    output = (subprocess.run(["docker", "run", "-it", "--rm", "--network", "host", "nfs01.techstudio.tv/ffprobe:latest", "-v", "quiet", "-print_format", "json", "-show_format", "-show_streams", "-i", mCast], stdout=subprocess.PIPE))


    data = json.loads(output.stdout.decode("utf8"))
    if not data:
        continue
    print(data)


    if format is "1":
        ws['A' + str(n)] = str(n-1)
        ws['C' + str(n)] = mCast
        sub_check = 0
        aud_check = 0
        for check in data["streams"]:
            if "codec_type" not in check:
                continue
            elif check["codec_type"] == "video":
                print(check["codec_name"])
                ws['E' + str(n)] = check["codec_name"]
                resolution = str(check["width"]) + "x" + str(check["height"])
            elif check["codec_type"] == "audio":
                if aud_check != 0:
                    print(check["codec_name"])
                    ws['G' + str(n)] = check["codec_name"]
                    aud_check = 0
                else:
                    print(check["codec_name"])
                    ws['F' + str(n)] = check["codec_name"]
                    aud_check += 1
            elif check["codec_type"] == "subtitle":
                if sub_check != 0:                  
                    ws['I' + str(n)] = check["codec_name"]
                    print(check["codec_name"])
                    sub_check = 0
                else:
                    ws['H' + str(n)] = check["codec_name"]
                    print(check["codec_name"])
                    sub_check += 1
            

        if not resolution:
            resolution = "--------"
        ws['D' + str(n)] = resolution
        n += 1
    else:
        json_f = open("m9.json", "a")
        json_f.write(json.dumps(data))
        json_f.close



if format is "1":
    wb.save('Channels.xlsx')
